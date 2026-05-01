#!/usr/bin/env python3
"""
Generate QC Excel spreadsheets for Chenwei to review new dataset collections.

Creates one Excel file per class with traceability fields:
  - #, Filename, SwatchOn ID, Source URL, Image URL
  - Fabric Type, Fiber Content, Pattern, Weight
  - Quality OK? (Y/N), Notes

Supports:
  - SwatchOn collections (Tricot, Ribbed_Poplin, Leno_Gauze) with JSON metadata
  - Interlock (Fabric Wholesale Direct, no JSON)
  - Woven+Jacquard (from relabel mapping CSV)

Usage:
  python scripts/generate_qc_excel.py                    # all classes
  python scripts/generate_qc_excel.py --only Tricot      # single class
  python scripts/generate_qc_excel.py --output-dir qc/   # custom output dir
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)


DATASET_ROOT = "FabricFlow_Dataset"
OUTPUT_DIR = "qc_excel"

# Class definitions
CLASSES = {
    "Tricot": {
        "l1": "KNIT",
        "source": "swatchon",
        "has_json": True,
    },
    "Ribbed_Poplin": {
        "l1": "WOVEN",
        "source": "swatchon",
        "has_json": True,
    },
    "Leno_Gauze": {
        "l1": "WOVEN",
        "source": "swatchon",
        "has_json": True,
    },
    "Interlock": {
        "l1": "KNIT",
        "source": "fabric_wholesale_direct",
        "has_json": False,
        "alt_dir": os.path.join("outputs", "fwd_interlock", "cropped_patches"),
    },
    "Woven+Jacquard": {
        "l1": "WOVEN",
        "source": "swatchon",
        "has_json": False,  # Uses CSV mapping instead
        "use_csv": True,
    },
}

# Column headers
HEADERS = [
    "#",
    "Filename",
    "Dataset Class",
    "L1 (Construction)",
    "Source",
    "SwatchOn ID",
    "Source URL",
    "Image URL",
    "Fabric Type",
    "Fiber Content",
    "Pattern",
    "Weight",
    "Width",
    "Dye Method",
    "Characteristics",
    "Country",
    "Quality OK? (Y/N)",
    "Notes",
]


def extract_swatchon_id_from_url(url: str) -> str:
    """Extract quality ID from SwatchOn URL."""
    m = re.search(r"-(\d+)$", url.rstrip("/"))
    return m.group(1) if m else ""


def load_swatchon_entries(class_name: str, cfg: dict) -> List[Dict[str, str]]:
    """Load entries from FabricFlow_Dataset with JSON metadata."""
    data_dir = os.path.join(DATASET_ROOT, cfg["l1"], class_name, cfg["source"])
    if not os.path.isdir(data_dir):
        print(f"  [warn] Directory not found: {data_dir}")
        return []

    jpgs = sorted([f for f in os.listdir(data_dir) if f.endswith(".jpg")])
    entries = []

    for idx, jpg in enumerate(jpgs, 1):
        json_path = os.path.join(data_dir, jpg.replace(".jpg", ".json"))
        row = {
            "#": str(idx),
            "Filename": jpg,
            "Dataset Class": class_name,
            "L1 (Construction)": cfg["l1"],
            "Source": "SwatchOn",
        }

        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as f:
                meta = json.load(f)
            specs = meta.get("specifications", {})
            row["SwatchOn ID"] = extract_swatchon_id_from_url(meta.get("detail_url", ""))
            row["Source URL"] = meta.get("detail_url", "")
            row["Image URL"] = meta.get("image_src", "")
            row["Fabric Type"] = specs.get("Fabric Type", "")
            row["Fiber Content"] = specs.get("Fiber Content", "")
            row["Pattern"] = specs.get("Pattern", "")
            row["Weight"] = specs.get("Weight", "")
            row["Width"] = specs.get("Width", "")
            row["Dye Method"] = specs.get("Dye Method", "")
            row["Characteristics"] = specs.get("Characteristics", "")
            row["Country"] = specs.get("Country", "")

        row["Quality OK? (Y/N)"] = ""
        row["Notes"] = ""
        entries.append(row)

    return entries


def load_interlock_entries(cfg: dict) -> List[Dict[str, str]]:
    """Load Interlock entries (no JSON metadata)."""
    data_dir = cfg.get("alt_dir", "")
    if not os.path.isdir(data_dir):
        print(f"  [warn] Directory not found: {data_dir}")
        return []

    jpgs = sorted([f for f in os.listdir(data_dir) if f.endswith(".jpg")])
    entries = []

    for idx, jpg in enumerate(jpgs, 1):
        row = {
            "#": str(idx),
            "Filename": jpg,
            "Dataset Class": "Interlock",
            "L1 (Construction)": "KNIT",
            "Source": "Fabric Wholesale Direct",
            "SwatchOn ID": "",
            "Source URL": "",
            "Image URL": "",
            "Fabric Type": "Interlock",
            "Fiber Content": "",
            "Pattern": "",
            "Weight": "",
            "Width": "",
            "Dye Method": "",
            "Characteristics": "",
            "Country": "",
            "Quality OK? (Y/N)": "",
            "Notes": "",
        }
        entries.append(row)

    return entries


def load_jacquard_entries() -> List[Dict[str, str]]:
    """Load Woven+Jacquard entries from relabel mapping CSV."""
    csv_path = "jacquard_relabel_mapping.csv"
    if not os.path.exists(csv_path):
        print(f"  [warn] CSV not found: {csv_path}")
        return []

    entries = []
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["original_folder"] != "Woven_Jacquard":
                continue
            quality_ok = row.get("quality_ok", "").strip()
            notes = row.get("notes", "").strip()

            m = re.match(r"QL-0*(\d+)_(\d+)", row.get("swatchon_id", ""))
            qid = m.group(1) if m else ""

            entry = {
                "#": str(len(entries) + 1),
                "Filename": row.get("new_filename", ""),
                "Dataset Class": "Woven+Jacquard",
                "L1 (Construction)": "WOVEN",
                "Source": "SwatchOn",
                "SwatchOn ID": row.get("swatchon_id", ""),
                "Source URL": f"https://swatchon.com/quality/{qid}" if qid else "",
                "Image URL": "",
                "Fabric Type": "Jacquard Weave",
                "Fiber Content": "",
                "Pattern": "",
                "Weight": "",
                "Width": "",
                "Dye Method": "",
                "Characteristics": "",
                "Country": "",
                "Quality OK? (Y/N)": quality_ok if quality_ok == "N" else "",
                "Notes": notes,
            }
            entries.append(entry)

    return entries


def create_excel(class_name: str, entries: List[Dict[str, str]], output_dir: str):
    """Create a styled QC Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{class_name} QC"

    # Styles
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    qc_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Write headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    for row_idx, entry in enumerate(entries, 2):
        for col_idx, header in enumerate(HEADERS, 1):
            val = entry.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # Highlight QC columns
            if header in ("Quality OK? (Y/N)", "Notes"):
                cell.fill = qc_fill

    # Column widths
    col_widths = {
        "#": 5, "Filename": 40, "Dataset Class": 16, "L1 (Construction)": 14,
        "Source": 14, "SwatchOn ID": 14, "Source URL": 45, "Image URL": 45,
        "Fabric Type": 20, "Fiber Content": 25, "Pattern": 14, "Weight": 10,
        "Width": 10, "Dye Method": 14, "Characteristics": 18, "Country": 8,
        "Quality OK? (Y/N)": 14, "Notes": 25,
    }
    for col_idx, header in enumerate(HEADERS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_widths.get(header, 12)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(HEADERS))}{len(entries) + 1}"

    # Save
    out_path = os.path.join(output_dir, f"{class_name}_QC.xlsx")
    wb.save(out_path)
    print(f"  Saved: {out_path} ({len(entries)} rows)")
    return out_path


def main():
    parser = argparse.ArgumentParser(description="Generate QC Excel files")
    parser.add_argument("--only", type=str, default="", help="Comma-separated class names")
    parser.add_argument("--output-dir", type=str, default=OUTPUT_DIR, help="Output directory")
    args = parser.parse_args()

    selected = {s.strip() for s in args.only.split(",") if s.strip()} if args.only else set()
    os.makedirs(args.output_dir, exist_ok=True)

    for class_name, cfg in CLASSES.items():
        if selected and class_name not in selected:
            continue

        print(f"\n{'='*50}")
        print(f"  {class_name}")
        print(f"{'='*50}")

        if cfg.get("use_csv"):
            entries = load_jacquard_entries()
        elif class_name == "Interlock":
            entries = load_interlock_entries(cfg)
        else:
            entries = load_swatchon_entries(class_name, cfg)

        if not entries:
            print(f"  No entries found, skipping.")
            continue

        create_excel(class_name, entries, args.output_dir)

    print(f"\nDone! QC files in: {args.output_dir}/")


if __name__ == "__main__":
    sys.exit(main() or 0)
