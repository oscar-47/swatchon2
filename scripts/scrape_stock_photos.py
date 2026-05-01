#!/usr/bin/env python3
"""
Scrape fabric texture images from Pexels and Unsplash for FabricFlow dataset.

Searches multiple queries per class, downloads images, deduplicates with dHash,
and saves with FabricFlow naming convention.

Usage:
  python scripts/scrape_stock_photos.py --class Cable_Knit
  python scripts/scrape_stock_photos.py --class Cable_Knit --target 300
  python scripts/scrape_stock_photos.py --class Cable_Knit --source pexels
  python scripts/scrape_stock_photos.py --class Cable_Knit --dry-run
  python scripts/scrape_stock_photos.py --all              # run all 6 classes
  python scripts/scrape_stock_photos.py --list-classes
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
import time
import urllib.error
import urllib.request
from typing import Any, Dict, List, Optional

# ── Config ──────────────────────────────────────────────────────────────

CONFIG_PATH = os.path.join("scripts", "config", "api_keys.json")

# Search queries per class (multiple queries increase diversity)
CLASS_QUERIES = {
    "Cable_Knit": {
        "l1": "KNIT",
        "pattern": "base",
        "target": 300,
        "queries": [
            "cable knit texture close up",
            "cable knit fabric pattern",
            "cable knit sweater texture",
            "aran knit pattern close up",
            "cable stitch knitting texture",
            "braided knit fabric",
        ],
        "sources": ["pexels", "unsplash"],
    },
    "Basket_Hopsack": {
        "l1": "WOVEN",
        "pattern": "base",
        "target": 300,
        "queries": [
            "basket weave fabric texture",
            "hopsack fabric close up",
            "basket weave textile",
            "panama weave fabric",
            "open weave fabric texture",
            "burlap basket weave close up",
        ],
        "sources": ["pexels", "unsplash"],
    },
    "Purl_Knit": {
        "l1": "KNIT",
        "pattern": "base",
        "target": 300,
        "queries": [
            "purl stitch knit texture",
            "reverse stockinette fabric",
            "purl knit pattern close up",
            "knitted purl texture",
            "seed stitch knit fabric",
            "garter stitch knit texture",
        ],
        "sources": ["pexels", "unsplash"],
    },
    "Double_Jersey": {
        "l1": "KNIT",
        "pattern": "base",
        "target": 400,
        "queries": [
            "double jersey fabric texture",
            "double knit fabric close up",
            "ponte fabric texture",
            "double face jersey textile",
            "double jersey knit",
            "scuba fabric texture close up",
        ],
        "sources": ["pexels", "unsplash"],
    },
    "Intarsia": {
        "l1": "KNIT",
        "pattern": "base",
        "target": 200,
        "queries": [
            "intarsia knit pattern",
            "intarsia knitting close up",
            "colorwork knit intarsia",
            "intarsia sweater pattern",
            "multi color knit intarsia",
        ],
        "sources": ["pexels", "unsplash"],
    },
    "Raschel": {
        "l1": "KNIT",
        "pattern": "base",
        "target": 200,
        "queries": [
            "raschel knit fabric",
            "raschel lace fabric texture",
            "raschel net fabric",
            "warp knit lace close up",
            "raschel curtain fabric",
        ],
        "sources": ["pexels", "unsplash"],
    },
}

# ── API helpers ─────────────────────────────────────────────────────────

def load_api_keys() -> Dict[str, str]:
    with open(CONFIG_PATH) as f:
        return json.load(f)


def search_pexels(query: str, api_key: str, page: int = 1, per_page: int = 80) -> List[Dict]:
    """Search Pexels, return list of {id, url, download_url, width, height, photographer}."""
    url = f"https://api.pexels.com/v1/search?query={urllib.request.quote(query)}&per_page={per_page}&page={page}"
    req = urllib.request.Request(url, headers={
        "Authorization": api_key,
        "User-Agent": "Mozilla/5.0",
    })
    try:
        resp = urllib.request.urlopen(req, timeout=20)
        data = json.loads(resp.read())
    except Exception as e:
        print(f"    Pexels error: {e}")
        return []

    results = []
    for p in data.get("photos", []):
        results.append({
            "id": f"pexels_{p['id']}",
            "source": "pexels",
            "url": p["url"],
            "download_url": p["src"]["original"],
            "width": p["width"],
            "height": p["height"],
            "photographer": p.get("photographer", ""),
            "alt": p.get("alt", ""),
        })
    return results


def search_unsplash(query: str, api_key: str, page: int = 1, per_page: int = 30) -> List[Dict]:
    """Search Unsplash, return list of {id, url, download_url, width, height, photographer}."""
    url = f"https://api.unsplash.com/search/photos?query={urllib.request.quote(query)}&per_page={per_page}&page={page}"
    req = urllib.request.Request(url, headers={
        "Authorization": f"Client-ID {api_key}",
        "Accept": "application/json",
    })
    try:
        resp = urllib.request.urlopen(req, timeout=20)
        data = json.loads(resp.read())
    except Exception as e:
        print(f"    Unsplash error: {e}")
        return []

    results = []
    for r in data.get("results", []):
        results.append({
            "id": f"unsplash_{r['id']}",
            "source": "unsplash",
            "url": r["links"]["html"],
            "download_url": r["urls"]["regular"],  # 1080px wide
            "width": r["width"],
            "height": r["height"],
            "photographer": r.get("user", {}).get("name", ""),
            "alt": r.get("alt_description", ""),
        })
    return results


# ── Image download & dedup ──────────────────────────────────────────────

def dhash(image_data: bytes, hash_size: int = 8) -> str:
    """Compute dHash from raw image bytes using PIL."""
    from PIL import Image
    import io
    img = Image.open(io.BytesIO(image_data)).convert("L").resize((hash_size + 1, hash_size), Image.LANCZOS)
    pixels = list(img.getdata())
    bits = []
    for row in range(hash_size):
        for col in range(hash_size):
            idx = row * (hash_size + 1) + col
            bits.append(1 if pixels[idx] < pixels[idx + 1] else 0)
    return "".join(str(b) for b in bits)


def hamming_distance(h1: str, h2: str) -> int:
    return sum(c1 != c2 for c1, c2 in zip(h1, h2))


def is_duplicate(new_hash: str, existing_hashes: set, threshold: int = 5) -> bool:
    for h in existing_hashes:
        if hamming_distance(new_hash, h) <= threshold:
            return True
    return False


def download_image(url: str, timeout: int = 30) -> Optional[bytes]:
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        resp = urllib.request.urlopen(req, timeout=timeout)
        return resp.read()
    except Exception:
        return None


# ── Main ────────────────────────────────────────────────────────────────

def scrape_class(class_name: str, cfg: dict, keys: dict, args):
    l1 = cfg["l1"]
    pattern = cfg["pattern"]
    target = args.target or cfg["target"]
    sources = [args.source] if args.source else cfg["sources"]

    # Output dir
    out_dir = os.path.join("FabricFlow_Dataset", l1, class_name)
    os.makedirs(out_dir, exist_ok=True)

    # Track existing files per source
    existing = {}
    for src in sources:
        src_dir = os.path.join(out_dir, src)
        os.makedirs(src_dir, exist_ok=True)
        existing[src] = len([f for f in os.listdir(src_dir) if f.endswith(".jpg")])

    print(f"\n{'='*60}")
    print(f"  {class_name} (target: {target})")
    print(f"  Sources: {sources}")
    print(f"  Existing: {existing}")
    print(f"{'='*60}")

    # Collect all search results
    all_results: Dict[str, Dict] = {}  # id -> metadata
    for query in cfg["queries"]:
        for src in sources:
            for page in range(1, 4):  # up to 3 pages per query
                if src == "pexels":
                    results = search_pexels(query, keys["pexels"], page=page)
                elif src == "unsplash":
                    results = search_unsplash(query, keys["unsplash"], page=page)
                else:
                    continue

                new_count = 0
                for r in results:
                    if r["id"] not in all_results:
                        all_results[r["id"]] = {**r, "query": query}
                        new_count += 1

                print(f"  [{src}] '{query}' p{page}: {len(results)} results, {new_count} new (total: {len(all_results)})")

                if len(results) == 0:
                    break
                time.sleep(0.3)

                if len(all_results) >= target * 2:
                    break
            if len(all_results) >= target * 2:
                break

    print(f"\n  Total candidates: {len(all_results)}")

    if args.dry_run:
        print(f"  (dry-run) Would download up to {target} images")
        return

    # Download and deduplicate
    hashes: set = set()
    downloaded = {src: existing.get(src, 0) for src in sources}
    total_downloaded = sum(downloaded.values())
    skipped = 0
    failed = 0

    for item in all_results.values():
        if total_downloaded >= target:
            break

        src = item["source"]
        src_dir = os.path.join(out_dir, src)
        seq = downloaded[src] + 1
        class_lower = class_name.lower()
        filename = f"{class_lower}_{pattern}_{src}_{seq:04d}.jpg"
        out_path = os.path.join(src_dir, filename)

        # Download
        img_data = download_image(item["download_url"])
        if not img_data:
            failed += 1
            continue

        # Check minimum size
        from PIL import Image
        import io
        try:
            img = Image.open(io.BytesIO(img_data))
            if img.width < 224 or img.height < 224:
                skipped += 1
                continue
        except Exception:
            failed += 1
            continue

        # Dedup
        h = dhash(img_data)
        if is_duplicate(h, hashes):
            skipped += 1
            continue
        hashes.add(h)

        # Save image
        with open(out_path, "wb") as f:
            f.write(img_data)

        # Save metadata JSON
        meta = {
            "source": src,
            "source_id": item["id"],
            "source_url": item["url"],
            "download_url": item["download_url"],
            "query": item["query"],
            "photographer": item["photographer"],
            "alt": item["alt"],
            "original_size": f"{item['width']}x{item['height']}",
        }
        json_path = out_path.replace(".jpg", ".json")
        with open(json_path, "w") as f:
            json.dump(meta, f, indent=2, ensure_ascii=False)

        downloaded[src] += 1
        total_downloaded += 1

        if total_downloaded % 25 == 0:
            print(f"  Progress: {total_downloaded}/{target} ({downloaded})")

        time.sleep(0.2)

    print(f"\n  Results:")
    print(f"    Downloaded: {total_downloaded} (by source: {downloaded})")
    print(f"    Skipped (dup/small): {skipped}")
    print(f"    Failed: {failed}")
    print(f"    Output: {out_dir}/")

    # Generate QC Excel
    generate_qc_excel(class_name, cfg, out_dir)


def generate_qc_excel(class_name: str, cfg: dict, out_dir: str):
    """Generate QC Excel from downloaded images + JSON metadata."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  [warn] openpyxl not installed, skipping QC Excel")
        return

    headers = [
        "#", "Filename", "Dataset Class", "L1 (Construction)", "Source",
        "Photo ID", "Source URL", "Photographer", "Search Query",
        "Description", "Original Size",
        "Quality OK? (Y/N)", "Notes",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{class_name} QC"

    # Styles
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    qc_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Collect all jpg+json pairs from all source subdirs
    rows = []
    for src_name in sorted(os.listdir(out_dir)):
        src_dir = os.path.join(out_dir, src_name)
        if not os.path.isdir(src_dir) or src_name.startswith("."):
            continue
        for jpg in sorted(f for f in os.listdir(src_dir) if f.endswith(".jpg")):
            json_path = os.path.join(src_dir, jpg.replace(".jpg", ".json"))
            meta = {}
            if os.path.exists(json_path):
                with open(json_path) as f:
                    meta = json.load(f)
            rows.append({
                "#": str(len(rows) + 1),
                "Filename": jpg,
                "Dataset Class": class_name,
                "L1 (Construction)": cfg["l1"],
                "Source": meta.get("source", src_name),
                "Photo ID": meta.get("source_id", ""),
                "Source URL": meta.get("source_url", ""),
                "Photographer": meta.get("photographer", ""),
                "Search Query": meta.get("query", ""),
                "Description": meta.get("alt", ""),
                "Original Size": meta.get("original_size", ""),
                "Quality OK? (Y/N)": "",
                "Notes": "",
            })

    for row_idx, row in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if h in ("Quality OK? (Y/N)", "Notes"):
                cell.fill = qc_fill

    # Column widths
    widths = {
        "#": 5, "Filename": 38, "Dataset Class": 16, "L1 (Construction)": 14,
        "Source": 10, "Photo ID": 18, "Source URL": 45, "Photographer": 20,
        "Search Query": 28, "Description": 35, "Original Size": 14,
        "Quality OK? (Y/N)": 14, "Notes": 25,
    }
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = widths.get(h, 12)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows) + 1}"

    os.makedirs("qc_excel", exist_ok=True)
    out_path = os.path.join("qc_excel", f"{class_name}_QC.xlsx")
    wb.save(out_path)
    print(f"\n  QC Excel: {out_path} ({len(rows)} rows)")


def main():
    parser = argparse.ArgumentParser(description="Scrape stock photos for FabricFlow")
    parser.add_argument("--class", dest="cls", type=str, help="Class name to scrape")
    parser.add_argument("--target", type=int, default=0, help="Override target count")
    parser.add_argument("--source", type=str, choices=["pexels", "unsplash"], help="Single source only")
    parser.add_argument("--dry-run", action="store_true", help="Search only, don't download")
    parser.add_argument("--all", action="store_true", help="Run all 6 classes")
    parser.add_argument("--list-classes", action="store_true", help="List available classes")
    args = parser.parse_args()

    if args.list_classes:
        for name, cfg in CLASS_QUERIES.items():
            print(f"  {name}: target={cfg['target']}, sources={cfg['sources']}")
        return

    keys = load_api_keys()
    if not keys.get("pexels"):
        print("Error: pexels key missing in scripts/config/api_keys.json")
        return 1

    if args.all:
        for name, cfg in CLASS_QUERIES.items():
            scrape_class(name, cfg, keys, args)
        return

    if not args.cls:
        print("Error: --class or --all required. Use --list-classes to see options.")
        return 1

    if args.cls not in CLASS_QUERIES:
        print(f"Error: unknown class '{args.cls}'. Available: {list(CLASS_QUERIES.keys())}")
        return 1

    scrape_class(args.cls, CLASS_QUERIES[args.cls], keys, args)


if __name__ == "__main__":
    sys.exit(main() or 0)
