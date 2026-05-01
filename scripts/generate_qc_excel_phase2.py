#!/usr/bin/env python3
"""
Generate QC Excel for Phase 2 collections (professional sources only).
One Excel per class. Reads JSON metadata if available.

Usage:
  python scripts/generate_qc_excel_phase2.py
  python scripts/generate_qc_excel_phase2.py --only Double_Jersey
"""
import argparse, json, os, sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

try:
    from PIL import Image
except ImportError:
    Image = None

DATASET_ROOT = Path("FabricFlow_Dataset")
OUTPUT_DIR = Path("qc_excel/phase2")

# Classes to generate QC for — professional sources only (skip pexels/unsplash)
SKIP_SOURCES = {"pexels", "unsplash"}

CLASSES = {
    "Double_Jersey":   "KNIT",
    "Cable_Knit":      "KNIT",
    "Purl_Knit":       "KNIT",
    "Intarsia":        "KNIT",
    "Raschel":         "KNIT",
    "Basket_Hopsack":  "WOVEN",
}

HEADERS = [
    "#",
    "Filename",
    "Source",
    "Class",
    "L1",
    "Resolution",
    "Source URL",
    "Product Name",
    "Quality OK? (Y/N)",
    "Notes",
]

# Styles
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
ALT_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def get_resolution(jpg_path):
    if Image is None:
        return ""
    try:
        with Image.open(jpg_path) as img:
            return f"{img.size[0]}x{img.size[1]}"
    except:
        return ""


def load_meta(json_path):
    """Load JSON metadata, return dict or empty dict."""
    if not json_path.exists():
        return {}
    try:
        with open(json_path) as f:
            return json.load(f)
    except:
        return {}


def generate_class_excel(class_name, l1):
    class_dir = DATASET_ROOT / l1 / class_name
    if not class_dir.exists():
        print(f"  [skip] {class_dir} not found")
        return 0

    # Collect all JPGs from professional sources
    rows = []
    for source_dir in sorted(class_dir.iterdir()):
        if not source_dir.is_dir():
            continue
        source = source_dir.name
        if source in SKIP_SOURCES:
            continue

        jpgs = sorted(source_dir.glob("*.jpg"))
        if not jpgs:
            continue

        for jpg in jpgs:
            json_path = jpg.with_suffix(".json")
            meta = load_meta(json_path)

            source_url = (
                meta.get("source_url") or
                meta.get("detail_url") or
                meta.get("page_url") or
                meta.get("image_url") or
                ""
            )
            product_name = (
                meta.get("item_name") or
                meta.get("product_title") or
                meta.get("alt") or
                ""
            )

            rows.append({
                "Filename": jpg.name,
                "Source": source,
                "Class": class_name,
                "L1": l1,
                "Resolution": get_resolution(jpg),
                "Source URL": source_url,
                "Product Name": product_name,
                "jpg_path": str(jpg),
            })

    if not rows:
        print(f"  [skip] {class_name}: no professional source images")
        return 0

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = class_name

    # Header row
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        values = [
            row_idx - 1,                    # #
            row_data["Filename"],           # Filename
            row_data["Source"],             # Source
            row_data["Class"],              # Class
            row_data["L1"],                 # L1
            row_data["Resolution"],         # Resolution
            row_data["Source URL"],          # Source URL
            row_data["Product Name"],       # Product Name
            "",                             # Quality OK?
            "",                             # Notes
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx >= 7))

        # Alternating row color
        if row_idx % 2 == 0:
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = ALT_FILL

    # Column widths
    col_widths = [5, 40, 18, 18, 8, 12, 50, 40, 14, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Class").font = Font(bold=True)
    ws2.cell(row=1, column=2, value=class_name).font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value="Total Images")
    ws2.cell(row=2, column=2, value=len(rows))
    ws2.cell(row=3, column=1, value="Generated")
    ws2.cell(row=3, column=2, value=str(__import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M")))

    # Per-source breakdown
    ws2.cell(row=5, column=1, value="Source").font = Font(bold=True)
    ws2.cell(row=5, column=2, value="Count").font = Font(bold=True)
    source_counts = {}
    for r in rows:
        source_counts[r["Source"]] = source_counts.get(r["Source"], 0) + 1
    for i, (src, cnt) in enumerate(sorted(source_counts.items()), 6):
        ws2.cell(row=i, column=1, value=src)
        ws2.cell(row=i, column=2, value=cnt)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 30

    # Save
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"{class_name}_QC.xlsx"
    wb.save(str(out_path))
    print(f"  ✓ {class_name}: {len(rows)} rows → {out_path}")
    return len(rows)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--only", help="Generate for single class only")
    args = parser.parse_args()

    classes = CLASSES
    if args.only:
        if args.only not in CLASSES:
            print(f"Unknown class: {args.only}. Available: {list(CLASSES.keys())}")
            sys.exit(1)
        classes = {args.only: CLASSES[args.only]}

    print(f"Generating QC Excel for {len(classes)} classes...")
    print(f"Output: {OUTPUT_DIR}/\n")

    total = 0
    for class_name, l1 in classes.items():
        n = generate_class_excel(class_name, l1)
        total += n

    print(f"\nDone. {total} total rows across {len(classes)} files.")


if __name__ == "__main__":
    main()
